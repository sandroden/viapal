# /// script
# dependencies = ["openpyxl"]
# ///
"""Estrae le schede 'conti bruna 2025' e 'Conti  Bruna 2024' da bruna.xlsx
e produce CSV puliti con (data, descrizione, importo_signed).

Convenzione importo:
- Foglio 2025: col B descrizione, col C importo (signed)
- Foglio 2024: col B descrizione, col E entrate, col F uscite -> signed

Uso:
  uv run scripts/dump_bruna_sheets.py -i dati/bruna-sheets/bruna.xlsx -o dati/
"""
import argparse
import csv
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path


def cell_to_iso(c) -> str | None:
    if c is None or c == "":
        return None
    if isinstance(c, datetime):
        return c.date().isoformat()
    if isinstance(c, date):
        return c.isoformat()
    try:
        return datetime.strptime(str(c), "%d/%m/%Y").date().isoformat()
    except ValueError:
        return None


def to_decimal(c):
    if c is None or c == "":
        return None
    try:
        return Decimal(str(c).replace(",", "."))
    except Exception:
        return None


def parse_2025(ws) -> list[dict]:
    out = []
    for row in ws.iter_rows(values_only=True):
        d = cell_to_iso(row[0])
        descr = (row[1] or "").strip() if row[1] else ""
        imp = to_decimal(row[2])
        if not d or not descr or imp is None:
            continue
        out.append({"data": d, "descrizione": descr, "importo_signed": str(imp)})
    return out


def parse_2024(ws) -> list[dict]:
    """Foglio 2024: col A data, col B descrizione, col E entrate, col F uscite."""
    out = []
    for row in ws.iter_rows(values_only=True):
        d = cell_to_iso(row[0])
        descr = (row[1] or "").strip() if row[1] else ""
        if not d:
            continue
        entrata = to_decimal(row[4]) if len(row) > 4 else None
        uscita = to_decimal(row[5]) if len(row) > 5 else None
        if entrata is not None and entrata != 0:
            out.append({"data": d, "descrizione": descr,
                        "importo_signed": str(entrata)})
        if uscita is not None and uscita != 0:
            out.append({"data": d, "descrizione": descr,
                        "importo_signed": str(-uscita)})
    return out


def main():
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("-i", "--input", required=True, help="bruna.xlsx")
    ap.add_argument("-o", "--out", required=True, help="cartella di output")
    args = ap.parse_args()

    import openpyxl
    wb = openpyxl.load_workbook(args.input, data_only=True)

    out_dir = Path(args.out)
    out_dir.mkdir(parents=True, exist_ok=True)

    cfg = [
        ("conti bruna 2025", "movimenti-bruna-2025.csv", parse_2025),
        ("Conti  Bruna 2024", "movimenti-bruna-2024.csv", parse_2024),
    ]
    for sheet_name, out_name, fn in cfg:
        ws = wb[sheet_name]
        rows = fn(ws)
        rows.sort(key=lambda r: r["data"])
        path = out_dir / out_name
        with open(path, "w", newline="") as f:
            w = csv.DictWriter(f, fieldnames=["data", "descrizione", "importo_signed"])
            w.writeheader()
            w.writerows(rows)
        print(f"  {sheet_name}: {len(rows)} righe -> {path}")


if __name__ == "__main__":
    main()
