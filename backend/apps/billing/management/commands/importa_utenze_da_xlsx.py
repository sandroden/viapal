"""Importa utenze pubblicate dal foglio di calcolo storico (ContiAffittoUtenze.xlsx).

Layout dei fogli (uno per anno):

    riga 1: header  | mese | presenze tot | luce | gas | TARI | tot | tot/pres | <inq1> | <inq2> | ...
    coppia di righe per ogni periodo:
        riga A: <nome periodo> + presenze tot + luce + gas + tari + tot + tot/pres + giorni-per-inquilino
        riga B: (vuoto/etichetta) + (...) + quote-€-per-inquilino

Il foglio è la fonte di verità: la quota in € della riga B finisce in
``Receivable.importo_dovuto`` senza ricalcoli.
"""
import calendar
import datetime as dt
import re
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction


# Mapping esplicito: nome (lowercase) nel foglio → nominativo esatto nel db.
# Usato come prima ricerca, prima di iexact/icontains su varianti.
TENANT_DB_NAME = {
    "mariasevera": "Maria Severa Armas",
    "maria severa": "Maria Severa Armas",
}

# Varianti accettate per la ricerca fallback (iexact / icontains).
TENANT_ALIASES = {
    "mariasevera": ["mariasevera", "maria severa"],
    "maria severa": ["mariasevera", "maria severa"],
}

# Nome mese italiano → numero
MESI = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4,
    "maggio": 5, "giugno": 6, "luglio": 7, "agosto": 8,
    "settembre": 9, "ottobre": 10, "novembre": 11, "dicembre": 12,
}

# Header colonne fisse
COLS_FISSE = ("mese", "presenze tot", "luce", "gas", "tari", "tot", "tot/pres")


class Command(BaseCommand):
    help = "Importa le utenze pubblicate dal foglio di calcolo storico (xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Percorso del file .xlsx")
        parser.add_argument("--anno", type=int, help="Importa solo il foglio dell'anno indicato.")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla, mostra solo il parsing.")
        parser.add_argument(
            "--crea-tenant-mancanti", action="store_true",
            help="Crea TenantProfile per nomi non presenti nel db (storici).",
        )
        parser.add_argument(
            "--property", type=str, default=None,
            help="Immobile (id o nome). Obbligatorio se ci sono più immobili.",
        )

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # noqa
            raise CommandError("openpyxl non installato: `uv add openpyxl`") from e

        from properties.context import resolve_property_cli

        try:
            prop = resolve_property_cli(opts.get("property"))
        except ValueError as e:
            raise CommandError(str(e)) from e

        xlsx_path = Path(opts["xlsx_path"]).expanduser().resolve()
        if not xlsx_path.exists():
            raise CommandError(f"File non trovato: {xlsx_path}")

        wb = load_workbook(xlsx_path, data_only=True)
        anno_filtro = opts.get("anno")
        dry_run = opts["dry_run"]
        crea_tenant = opts["crea_tenant_mancanti"]

        report = []
        with transaction.atomic():
            for sheet_name in wb.sheetnames:
                if not re.fullmatch(r"\d{4}", sheet_name):
                    continue
                anno = int(sheet_name)
                if anno_filtro and anno != anno_filtro:
                    continue
                summary = self._import_sheet(
                    wb[sheet_name], anno, dry_run=dry_run, crea_tenant=crea_tenant,
                    prop=prop,
                )
                report.append(summary)

            if dry_run:
                self.stdout.write(self.style.WARNING("DRY-RUN: rollback transazione."))
                transaction.set_rollback(True)

        self._stampa_report(report)

    # ------------------------------------------------------------------
    # parser di un foglio
    # ------------------------------------------------------------------
    def _import_sheet(self, ws, anno: int, dry_run: bool, crea_tenant: bool, prop) -> dict:
        from billing.models import Receivable, UtilityChargePeriod
        from properties.models import RoomAssignment, TenantProfile

        header = [self._norm(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        # mappa nome inquilino -> indice colonna (tutto ciò che non è in COLS_FISSE)
        inq_cols: dict[str, int] = {}
        for idx, name in enumerate(header):
            if name and name not in COLS_FISSE:
                inq_cols[name] = idx

        rows = list(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True))

        periodi_creati = 0
        receivable_creati = 0
        receivable_aggiornati = 0
        tenant_mancanti: set[str] = set()
        assignment_mancanti: list[str] = []
        warnings_list: list[str] = []

        # Itera in coppie riga-A (giorni) / riga-B (€). Riga A = ha valore in col[0] (nome periodo).
        i = 0
        while i < len(rows):
            row_a = rows[i]
            if not row_a or row_a[0] is None or not isinstance(row_a[0], str):
                i += 1
                continue

            periodo_label = row_a[0].strip()
            periodo_da, periodo_a = self._parse_periodo(periodo_label, anno)
            if periodo_da is None:
                i += 1
                continue

            # riga B = riga immediatamente successiva con almeno una quota numerica
            row_b = rows[i + 1] if i + 1 < len(rows) else None
            if row_b is None:
                i += 1
                continue

            tot_luce = self._dec(row_a[2])
            tot_gas = self._dec(row_a[3])
            tot_tari = self._dec(row_a[4])
            giorni_totali_letti = self._int(row_a[1])

            # se mancano sia luce che gas, salto (regola Sandro: niente conguagli "solo TARI")
            if (tot_luce or 0) == 0 and (tot_gas or 0) == 0:
                warnings_list.append(
                    f"{anno} {periodo_label}: niente luce/gas, periodo saltato"
                )
                i += 2
                continue

            # ricalcolo giorni_totali dalla somma effettiva delle presenze (riga A),
            # robusto rispetto a inquilini omessi/aggiunti
            giorni_per_inq: dict[str, int] = {}
            quote_per_inq: dict[str, Decimal] = {}
            for nome, col_idx in inq_cols.items():
                g = self._int(row_a[col_idx])
                q = self._dec(row_b[col_idx])
                if g and g > 0:
                    giorni_per_inq[nome] = g
                if q is not None and q > 0:
                    quote_per_inq[nome] = q

            giorni_totali = sum(giorni_per_inq.values())
            if giorni_totali_letti and giorni_totali_letti != giorni_totali:
                warnings_list.append(
                    f"{anno} {periodo_label}: giorni_totali letto={giorni_totali_letti} "
                    f"≠ somma colonne={giorni_totali} (uso somma colonne)"
                )

            if dry_run:
                periodi_creati += 1
                receivable_creati += len(quote_per_inq)
                self.stdout.write(
                    f"  [{anno}] {periodo_label} ({periodo_da}→{periodo_a}): "
                    f"luce={tot_luce} gas={tot_gas} tari={tot_tari} g_tot={giorni_totali} "
                    f"inq={len(quote_per_inq)}"
                )
                i += 2
                continue

            # ---- persist ----
            period, created = UtilityChargePeriod.objects.update_or_create(
                property=prop,
                periodo_da=periodo_da,
                periodo_a=periodo_a,
                defaults={
                    "tot_luce": tot_luce or Decimal("0"),
                    "tot_gas": tot_gas or Decimal("0"),
                    "tot_tari": tot_tari or Decimal("0"),
                    "tot_altro": Decimal("0"),
                    "giorni_totali": giorni_totali,
                    "stato": "inviato",
                    "criterio_ripartizione": "pro_rata_giorni",
                },
            )
            if created:
                periodi_creati += 1

            scadenza = self._scadenza_periodo(periodo_a)

            for nome, quota in quote_per_inq.items():
                tenant = self._find_tenant(nome, crea_tenant=crea_tenant, prop=prop)
                if tenant is None:
                    tenant_mancanti.add(nome)
                    continue

                assignment = self._find_assignment(tenant, periodo_da, periodo_a)
                if assignment is None:
                    assignment_mancanti.append(
                        f"{anno} {periodo_label}: {tenant.nominativo} "
                        f"({giorni_per_inq.get(nome, '?')} giorni)"
                    )
                    continue

                _, r_created = Receivable.objects.update_or_create(
                    utility_period=period,
                    assignment=assignment,
                    causale=Receivable.Causale.UTENZE,
                    defaults={
                        "competenza_da": periodo_da,
                        "competenza_a": periodo_a,
                        "importo_dovuto": quota,
                        "giorni_presenza": giorni_per_inq.get(nome),
                        "scadenza": scadenza,
                    },
                )
                if r_created:
                    receivable_creati += 1
                else:
                    receivable_aggiornati += 1

            i += 2

        return {
            "anno": anno,
            "periodi_creati": periodi_creati,
            "receivable_creati": receivable_creati,
            "receivable_aggiornati": receivable_aggiornati,
            "tenant_mancanti": sorted(tenant_mancanti),
            "assignment_mancanti": assignment_mancanti,
            "warnings": warnings_list,
        }

    # ------------------------------------------------------------------
    # helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _norm(v) -> str:
        if v is None:
            return ""
        return str(v).strip().lower()

    @staticmethod
    def _dec(v) -> Decimal | None:
        if v is None or v == "" or isinstance(v, str) and "DIV/0" in v:
            return None
        try:
            return Decimal(str(v)).quantize(Decimal("0.01"))
        except Exception:
            return None

    @staticmethod
    def _int(v) -> int | None:
        if v is None or v == "" or isinstance(v, str) and "DIV/0" in v:
            return None
        try:
            return int(float(v))
        except Exception:
            return None

    def _parse_periodo(self, label: str, anno: int) -> tuple[dt.date, dt.date] | tuple[None, None]:
        """Estrae (periodo_da, periodo_a) da un'etichetta tipo 'Gennaio', 'Marzo-Aprile',
        'Settembre-Ottobre simulato'. Ritorna (None, None) se non parsabile.
        """
        normal = label.lower().strip()
        # rimuovi suffissi tipo "simulato"
        normal = re.sub(r"\bsimulato\b", "", normal).strip()

        # caso "mese1-mese2" o "mese1/mese2"
        m_range = re.match(r"^([a-z]+)\s*[-/]\s*([a-z]+)$", normal)
        if m_range:
            m1 = MESI.get(m_range.group(1))
            m2 = MESI.get(m_range.group(2))
            if m1 and m2:
                last_day = calendar.monthrange(anno, m2)[1]
                return dt.date(anno, m1, 1), dt.date(anno, m2, last_day)

        # caso singolo mese
        m_single = re.match(r"^([a-z]+)$", normal)
        if m_single and m_single.group(1) in MESI:
            m = MESI[m_single.group(1)]
            last_day = calendar.monthrange(anno, m)[1]
            return dt.date(anno, m, 1), dt.date(anno, m, last_day)

        return None, None

    @staticmethod
    def _scadenza_periodo(periodo_a: dt.date) -> dt.date:
        """5 giorni dopo la fine del periodo (in linea con il flusso 'pubblico a metà del mese
        successivo, scadenza ~ 5 del mese dopo')."""
        return periodo_a + dt.timedelta(days=5)

    @staticmethod
    def _find_tenant(nominativo_foglio: str, *, crea_tenant: bool, prop):
        """Cerca TenantProfile per nominativo (case-insensitive, alias gestiti)."""
        from django.contrib.auth.models import User

        from properties.models import TenantProfile

        nome_norm = nominativo_foglio.strip().lower()
        base_qs = TenantProfile.objects.filter(property=prop)

        # 1) mapping esplicito (più affidabile, preferito su match euristici)
        nome_db = TENANT_DB_NAME.get(nome_norm)
        if nome_db:
            t = base_qs.filter(nominativo__iexact=nome_db).first()
            if t:
                return t

        # 2) iexact sui candidati (alias o nome originale)
        candidati = TENANT_ALIASES.get(nome_norm, [nome_norm])
        for cand in candidati:
            t = base_qs.filter(nominativo__iexact=cand).first()
            if t:
                return t

        # 3) fallback icontains: se ambiguo (più di un match) → None per non
        # scegliere a caso, l'utente vedrà il warning e aggiungerà l'alias
        # in TENANT_DB_NAME.
        for cand in candidati:
            qs = base_qs.filter(nominativo__icontains=cand)
            if qs.count() == 1:
                return qs.first()

        if not crea_tenant:
            return None

        # crea User + TenantProfile storico
        username = f"storico_{nome_norm.replace(' ', '_')}"
        suffix = 0
        while User.objects.filter(username=username).exists():
            suffix += 1
            username = f"storico_{nome_norm.replace(' ', '_')}_{suffix}"
        user = User.objects.create_user(
            username=username,
            email=f"{username}@storico.local",
            password=None,
        )
        user.set_unusable_password()
        user.is_active = False
        user.save()
        return TenantProfile.objects.create(
            user=user,
            property=prop,
            nominativo=nominativo_foglio.title(),
            giorno_pagamento_affitto=1,
            note_pagamento="Inquilino storico, importato da ContiAffittoUtenze.xlsx",
        )

    @staticmethod
    def _find_assignment(tenant, periodo_da, periodo_a):
        """RoomAssignment del tenant attivo (anche solo parzialmente) nel periodo."""
        from django.db.models import Q

        from properties.models import RoomAssignment

        return (
            RoomAssignment.objects.filter(tenant=tenant, valid_from__lte=periodo_a)
            .filter(Q(valid_to__isnull=True) | Q(valid_to__gte=periodo_da))
            .order_by("-valid_from")
            .first()
        )

    # ------------------------------------------------------------------
    # report
    # ------------------------------------------------------------------
    def _stampa_report(self, report: list[dict]):
        for r in report:
            anno = r["anno"]
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(f"=== {anno} ==="))
            self.stdout.write(
                f"  periodi creati/aggiornati: {r['periodi_creati']}, "
                f"receivable creati: {r['receivable_creati']}, "
                f"aggiornati: {r['receivable_aggiornati']}"
            )
            for w in r["warnings"]:
                self.stdout.write(self.style.WARNING(f"  ⚠ {w}"))
            if r["tenant_mancanti"]:
                self.stdout.write(self.style.WARNING(
                    f"  inquilini non trovati nel db (usa --crea-tenant-mancanti): "
                    f"{', '.join(r['tenant_mancanti'])}"
                ))
            if r["assignment_mancanti"]:
                self.stdout.write(self.style.WARNING(
                    f"  RoomAssignment mancanti per {len(r['assignment_mancanti'])} righe:"
                ))
                for a in r["assignment_mancanti"][:20]:
                    self.stdout.write(f"      - {a}")
                if len(r["assignment_mancanti"]) > 20:
                    self.stdout.write(f"      ... (+{len(r['assignment_mancanti']) - 20})")
