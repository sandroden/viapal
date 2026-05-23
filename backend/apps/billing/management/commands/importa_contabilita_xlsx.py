"""Importa le spese dei proprietari dal libro mano (Contabilità.xlsx, tab «conti 2023»).

Questa prima fase importa SOLO le uscite dei proprietari, ovvero le voci a
rischio nullo di sovrapposizione con i Receivable già a sistema:

* ``tipo = a``  → arredo / manutenzione / pratiche immobile → ``Expense``
* ``tipo = g`` con "IMU" in descrizione → ``Expense`` (anticipata da un
  proprietario per conto di un altro, vedi ``riferimento_quota_owner``)

Tutti gli altri tipi (x, c, b, cauzione, p, altri g) vengono CONTATI e
riportati ma non scritti: l'affitto, le utenze e le cauzioni richiedono
l'accorpamento canone+condominio e la riconciliazione con i Receivable
esistenti — fase successiva.

Idempotente: ogni Expense importata porta nel campo ``note`` il marcatore
``[conti2023#NN]`` (NN = numero di riga nel foglio). Una seconda esecuzione
aggiorna le righe già viste invece di duplicarle. Le cifre del foglio sono
copiate senza arrotondamenti.
"""
import calendar
import datetime as dt
import re
from decimal import Decimal
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError
from django.db import transaction

# Codici tipo della colonna I del foglio.
TIPO_ESCLUSI = {"p"}            # personali — estranei all'affitto
TIPO_FASE2 = {"x", "c", "b", "cauzione"}  # affitto/utenze/cauzioni — fase successiva

# Colonna uscita (indice 0-based) → keyword per individuare l'OwnerProfile.
COL_USCITA_OWNER = {3: "Alessandro", 5: "Bruna", 7: "Fabio"}
COL_ENTRATA = (2, 4, 6)

# Marcatori di idempotenza nel campo note dei Receivable di fase 2.
MARKER_AFFITTO = "[conti2023-affitto]"
MARKER_CAUZIONE = "[conti2023-cauzione]"

# --- Fase 2: ricostruzione affitti mancanti di Maria Severa e Teo ---------
# Dovuto mensile = canone + quota condominio, letti dal foglio «conti 2023»
# (il foglio fa fede). Curato a mano riga per riga: affitto storico una tantum.
#   (tenant, anno, mese, dovuto, fonte righe foglio, nota)
AFFITTO_RICOSTRUITO = [
    ("Maria Severa", 2022, 10, "450", "riga 8",      "informale · 380 affitto + 70 condominio"),
    ("Maria Severa", 2022, 11, "450", "riga 5",      "informale · 380 + 70"),
    ("Maria Severa", 2022, 12, "450", "righe 26+27", "informale · 180+200 affitto + 70 condominio"),
    ("Maria Severa", 2023, 1,  "450", "riga 53",     "informale · 380 + 70 (pagato 28/01)"),
    ("Maria Severa", 2023, 2,  "450", "riga 35",     "informale · 380 + 70 (pagato 31/01, competenza febbraio)"),
    ("Maria Severa", 2023, 3,  "510", "righe 66+67", "transizione · 430 + 80 (condominio con TARI)"),
    ("Maria Severa", 2023, 4,  "510", "righe 79+80", "contrattualizzato · 430 + 80"),
    ("Maria Severa", 2023, 5,  "510", "righe 85+86", "contrattualizzato · 430 + 80"),
    ("Maria Severa", 2023, 6,  "510", "righe 101+102", "contrattualizzato · 430 + 80"),
    ("Maria Severa", 2023, 7,  "510", "righe 122+123", "contrattualizzato · 430 + 80"),
    ("Maria Severa", 2023, 8,  "510", "righe 131+132", "contrattualizzato · 430 + 80"),
    ("Teo",          2023, 4,  "600", "righe 82+84", "contrattualizzato · 530 + 70"),
    ("Teo",          2023, 5,  "600", "righe 92+93", "contrattualizzato · 530 + 70"),
    ("Teo",          2023, 6,  "600", "—",           "MESE ASSENTE DAL FOGLIO: ricostruito da canone 530 + 70 — DA VERIFICARE"),
    ("Teo",          2023, 7,  "600", "righe 124+125", "contrattualizzato · 530 + 70"),
    ("Teo",          2023, 8,  "600", "righe 133+134", "contrattualizzato · 530 + 70"),
    ("Alessandra",   2022, 10, "200", "riga 13",     "informale · 165 affitto + 35 condominio (mezzo mese, ingresso 15/10)"),
    ("Alessandra",   2022, 11, "400", "riga 11",     "informale · 330 + 70"),
    ("Alessandra",   2022, 12, "400", "riga 44",     "informale · 330 + 70 (pagato 13/01/2023)"),
]

# --- Fase 2: correzione cauzioni (il foglio fa fede) ----------------------
#   (tenant, importo_corretto, nota)
CAUZIONI_CORREZIONI = [
    ("Eugenia",  "530", "foglio riga 138"),
    ("Marianna", "600", "foglio righe 187+190 (acconto 400 + 200)"),
]

# --- Fase 2: correzione affitti già a sistema (il foglio fa fede) ----------
# I Receivable affitto Severa set-dic 2023 sono a 520 ma il foglio dice 500
# (430 + 70, c'è accordo di riduzione rispetto al canone ufficiale).
#   (tenant, anno, mese, importo_corretto, nota)
AFFITTO_CORREZIONI = [
    ("Maria Severa", 2023, 9,  "500", "foglio: 430 + 70 (accordo riduzione vs canone ufficiale)"),
    ("Maria Severa", 2023, 10, "500", "foglio: 430 + 70"),
    ("Maria Severa", 2023, 11, "500", "foglio: 430 + 70"),
    ("Maria Severa", 2023, 12, "500", "foglio: 430 + 70"),
]

# --- Fase 3: riconciliazione storica dal libro mano -----------------------
# I Receivable affitto/deposito 2022-2023 restano `atteso` perché mancano le
# BankTransaction. Sandro le ha dall'estratto conto reale (già importate);
# Bruna no (estratto conto non disponibile) e diversi incassi 2022 sono
# avvenuti in contanti. Qui le BankTransaction mancanti vengono *sintetizzate
# dal libro mano* «conti 2023» e allocate ai Receivable, così passano a
# `pagato`. Marcatore di idempotenza `[conti2023-bt …]` nel campo `note`
# della BT: una seconda esecuzione aggiorna invece di duplicare.
#
# Riga: (tenant_key, anno, mese, tranches, fonte, nota)
#   tranche = (owner_key, data_iso, importo) — un singolo movimento bancario.
#   Più tranche = pagamento spezzato su conti diversi (es. Severa dic 2022,
#   incassato in parte da Sandro e in parte da Fabio).
# L'importo è canone + condominio come li registra il foglio: se diverge dal
# dovuto del Receivable lo scarto viene *segnalato* (lo verifica Bruna), mai
# corretto in automatico.
MARKER_BT = "[conti2023-bt"
MARKER_CORR_COND = "[conti2023-corr-condominio]"

AFFITTO_FASE3 = [
    ("Severa", 2022, 10, [("Bruna", "2022-11-01", "450")], "righe 8+9", ""),
    ("Severa", 2022, 11, [("Alessandro", "2022-11-01", "450")], "righe 5+10",
     "incasso in contanti da Sandro"),
    ("Severa", 2022, 12, [("Alessandro", "2022-11-27", "250"),
                          ("Fabio", "2022-11-27", "200")], "righe 25+26+27",
     "incasso in contanti spezzato: Sandro 180+70 / Fabio 200"),
    ("Severa", 2023, 1, [("Alessandro", "2023-01-28", "450")], "righe 53+54",
     "incasso in contanti da Sandro"),
    ("Severa", 2023, 2, [("Bruna", "2023-01-31", "450")], "righe 34+35", ""),
    ("Severa", 2023, 3, [("Bruna", "2023-03-05", "510")], "righe 66+67", ""),
    ("Severa", 2023, 4, [("Bruna", "2023-04-03", "510")], "righe 79+80", ""),
    ("Severa", 2023, 5, [("Bruna", "2023-05-01", "510")], "righe 85+86", ""),
    ("Severa", 2023, 6, [("Bruna", "2023-06-04", "510")], "righe 101+102", ""),
    ("Severa", 2023, 7, [("Bruna", "2023-07-04", "510")], "righe 122+123", ""),
    ("Severa", 2023, 8, [("Bruna", "2023-08-07", "510")], "righe 131+132", ""),
    ("Severa", 2023, 9, [("Bruna", "2023-09-15", "500")], "righe 161+162", ""),
    ("Severa", 2023, 10, [("Bruna", "2023-10-03", "500")], "righe 173+174", ""),
    ("Severa", 2023, 11, [("Bruna", "2023-11-03", "500")], "righe 188+189", ""),
    ("Severa", 2023, 12, [("Bruna", "2023-12-04", "500")], "righe 196+198/199/200",
     "condominio = una delle 3 righe identiche 198-200"),
    ("Alessandra", 2022, 10, [("Bruna", "2022-10-15", "200")], "righe 12+13",
     "mezzo mese (ingresso 15/10)"),
    ("Alessandra", 2022, 11, [("Alessandro", "2022-11-01", "330")], "riga 11",
     "incasso in contanti da Sandro; condominio 70 non presente nel foglio"),
    ("Alessandra", 2022, 12, [("Bruna", "2023-01-13", "400")], "righe 44+45",
     "pagato 13/01/2023"),
    ("Teo", 2023, 4, [("Bruna", "2023-04-06", "600")], "righe 82+84", ""),
    ("Teo", 2023, 5, [("Bruna", "2023-05-07", "600")], "righe 92+93", ""),
    ("Teo", 2023, 6, [], "—", "mese assente dal libro mano: nessun movimento"),
    ("Teo", 2023, 7, [("Bruna", "2023-07-04", "600")], "righe 124+125", ""),
    ("Teo", 2023, 8, [("Bruna", "2023-08-07", "600")], "righe 133+134", ""),
    ("Eugenia", 2023, 9, [("Bruna", "2023-08-11", "530")], "riga 137",
     "condominio non presente nel foglio"),
    ("Eugenia", 2023, 10, [("Bruna", "2023-09-26", "670")], "righe 167+168",
     "condominio anomalo 140 nel foglio (atteso 70)"),
    ("Eugenia", 2023, 11, [("Bruna", "2023-11-06", "600")], "righe 191+192", ""),
    ("Eugenia", 2023, 12, [("Bruna", "2023-12-04", "600")], "righe 195+198/199/200",
     "condominio = una delle 3 righe identiche 198-200"),
    ("Marianna", 2023, 10, [("Bruna", "2023-09-20", "600")], "righe 164+165", ""),
    ("Marianna", 2023, 11, [("Bruna", "2023-11-02", "530")], "riga 186",
     "condominio non presente nel foglio"),
    ("Marianna", 2023, 12, [("Bruna", "2023-12-05", "600")], "righe 197+198/199/200",
     "condominio = una delle 3 righe identiche 198-200"),
]

# Depositi versati (positivi) ancora `atteso`. Le restituzioni cauzione
# (Receivable negativi) restano da riconciliare a mano: il foglio
# le registra in modo aggregato e non mappa 1:1 sui Receivable pro-rata.
#   (tenant_key, tranches, fonte, nota)
DEPOSITO_FASE3 = [
    ("Junior", [("Bruna", "2023-02-21", "1170")], "riga 61", ""),
    ("Teo", [("Bruna", "2023-04-06", "530")], "riga 83", ""),
    ("Eugenia", [("Bruna", "2023-08-11", "530")], "riga 138", ""),
    ("Marianna", [("Bruna", "2023-11-06", "600")], "righe 187+190",
     "acconto 400 + saldo 200"),
]

# Utenze 2023 di Maria Severa pagate in contanti (a Sandro o Bruna), tracciate
# solo sul libro mano. I Receivable causale=`utenze` esistono e sono `atteso`
# perché non esiste un movimento bancario reale: qui sintetizziamo la BT dal
# libro mano e la allochiamo (stesso meccanismo di AFFITTO_FASE3).
# Gli importi del foglio non sempre combaciano col dovuto pro-rata calcolato in
# app — gli scostamenti vengono segnalati ma non corretti (vince il foglio).
#   (tenant_key, anno, mese_competenza_da, tranches, fonte, nota)
UTENZE_FASE3 = [
    ("Severa", 2023, 2,
     [("Alessandro", "2023-01-28", "90")], "riga 55",
     "Utenze Maria Severa — incasso cash Sandro; 90€ del foglio > 51,24€ pro-rata"),
    ("Severa", 2023, 3,
     [("Bruna", "2023-06-06", "79")], "riga 104",
     "rimborso bollette severa — manca ~26,53€ rispetto al dovuto pro-rata"),
    ("Severa", 2023, 5,
     [("Bruna", "2023-09-05", "85.53")], "riga 158",
     "bollette Severa + tari — manca ~20€ rispetto al dovuto pro-rata"),
    ("Severa", 2023, 7,
     [("Bruna", "2023-10-03", "93")], "riga 175",
     "luce, gas e Tari Severa — importo combacia"),
    ("Severa", 2023, 9,
     [("Bruna", "2023-12-05", "73.81")], "riga 203",
     "utenze Severa — importo combacia"),
]

# Scarti di 1€ tra le Expense rate condominio già a sistema e il libro mano
# (vince il foglio). Match sulla descrizione *esatta* — esistono più "Nª rata"
# su cicli condominiali diversi.  (descrizione, importo_atteso, corretto, fonte)
CORREZIONI_CONDOMINIO = [
    ("Condominio 2022/2023 — 3ª rata MAV", "1246", "1247", "foglio riga 73"),
    ("Condominio 2022/2023 — 4ª rata MAV", "1209", "1210", "foglio riga 91"),
    ("Condominio 2023/2024 — 1ª rata MAV", "1209", "1210", "foglio riga 210"),
]

# --- Fase 4: bollette utenze ----------------------------------------------
# I PDF Edison `utenze-*edison.pdf` diventano UtilityBill (la cui post_save
# genera l'Expense lato proprietario). I metadati primari vengono dal nome
# file (`utenze-…-AAAA-MM-IMPORTO-edison.pdf`, prodotto + periodo + importo),
# arricchiti dal testo del PDF quando leggibile (i 2 PDF 2022 sono scansioni).
# La TARI (riga 121 del foglio) è una voce a sé → Expense categoria «tari».
MARKER_TARI = "[conti2023-tari]"

# Mesi italiani estesi — per il parsing dell'intestazione delle bollette Edison.
MESI_IT = {
    "gennaio": 1, "febbraio": 2, "marzo": 3, "aprile": 4, "maggio": 5,
    "giugno": 6, "luglio": 7, "agosto": 8, "settembre": 9, "ottobre": 10,
    "novembre": 11, "dicembre": 12,
}

# Bollette pagate da Bruna prima della voltura del contratto utenze a nome di
# Sandro: la fattura non era intestata a lui, quindi non esiste un PDF. Vanno
# a `UtilityBill` SENZA `pagata_da_owner` — così il signal non genera una
# Expense lato proprietario (come da indicazione: «come UtilityBill, non come
# Spese»). Il prodotto (luce/gas) non è determinabile dal libro mano: il foglio
# le aggrega come «bollette». Marcatore = numero_fattura `conti2023-bolletta-rigaNN`.
#   (riga, periodo_da, periodo_a, data_pagamento, importo, descr_foglio)
BOLLETTE_FOGLIO = [
    (81,  "2022-12-01", "2023-01-31", "2023-04-05", "208", "pagamento bollette dic/gen"),
    (88,  "2023-02-01", "2023-02-28", "2023-04-15", "76",  "pagamento bollette febbraio"),
    (103, "2023-03-01", "2023-04-30", "2023-06-05", "250", "pagamento bollette mar/apr"),
]


class Command(BaseCommand):
    help = "Importa le spese proprietari (tipo a + IMU) dal tab «conti 2023» di Contabilità.xlsx."

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str, help="Percorso del file .xlsx")
        parser.add_argument("--tab", default="conti 2023", help="Nome del foglio (default: «conti 2023»).")
        parser.add_argument("--max-row", type=int, default=214,
                            help="Ultima riga del foglio da importare (default: 214).")
        parser.add_argument("--fase", choices=["1", "2", "3", "4", "all"], default="all",
                            help="1 = spese proprietari; 2 = affitti + cauzioni "
                                 "ricostruiti; 3 = riconciliazione storica dal "
                                 "libro mano (BT sintetiche + correzioni); "
                                 "4 = bollette (PDF Edison → UtilityBill + TARI); "
                                 "all = tutte.")
        parser.add_argument("--bollette-dir", default=None,
                            help="Cartella dei PDF utenze-*edison.pdf per la fase 4 "
                                 "(default: la cartella del file .xlsx).")
        parser.add_argument("--dry-run", action="store_true", help="Non scrive nulla, mostra solo il parsing.")

    def handle(self, *args, **opts):
        try:
            from openpyxl import load_workbook
        except ImportError as e:  # noqa
            raise CommandError("openpyxl non installato: `uv add openpyxl`") from e

        xlsx_path = Path(opts["xlsx_path"]).expanduser().resolve()
        if not xlsx_path.exists():
            raise CommandError(f"File non trovato: {xlsx_path}")

        wb = load_workbook(xlsx_path, data_only=True)
        tab = opts["tab"]
        if tab not in wb.sheetnames:
            raise CommandError(f"Tab «{tab}» non trovato. Disponibili: {wb.sheetnames}")
        dry_run = opts["dry_run"]

        fase = opts["fase"]
        bollette_dir = opts["bollette_dir"] or str(xlsx_path.parent)
        summary1 = summary2 = summary3 = summary4 = None
        with transaction.atomic():
            if fase in ("1", "all"):
                summary1 = self._import(wb[tab], max_row=opts["max_row"], dry_run=dry_run)
            if fase in ("2", "all"):
                summary2 = self._fase2(dry_run=dry_run)
            if fase in ("3", "all"):
                summary3 = self._fase3(dry_run=dry_run)
            if fase in ("4", "all"):
                summary4 = self._fase4(bollette_dir=bollette_dir, dry_run=dry_run)
            if dry_run:
                self.stdout.write(self.style.WARNING("DRY-RUN: rollback transazione."))
                transaction.set_rollback(True)

        if summary1:
            self._report(summary1)
        if summary2:
            self._report_fase2(summary2)
        if summary3:
            self._report_fase3(summary3)
        if summary4:
            self._report_fase4(summary4)

    # ------------------------------------------------------------------
    def _import(self, ws, *, max_row: int, dry_run: bool) -> dict:
        from billing.models import Expense, ExpenseCategory
        from properties.models import OwnerProfile

        cat_arredo = self._categoria("arredo-manutenzione", "Arredo e manutenzione", dry_run)
        cat_imu = self._categoria("imu", "IMU", dry_run)
        owners = {kw: self._owner(kw) for kw in {"Alessandro", "Bruna", "Fabio"}}

        creati = aggiornati = 0
        tot_importato = Decimal("0")
        skip_buckets: dict[str, int] = {}
        warnings: list[str] = []
        ultima_data: dt.date | None = None

        for n, row in enumerate(ws.iter_rows(min_row=2, max_row=max_row, values_only=True), start=2):
            if not any(c not in (None, "") for c in row[:9]):
                continue
            data_raw, descr, tipo = row[0], row[1], row[8]
            descr = str(descr).strip() if descr is not None else ""
            tipo = str(tipo).strip() if tipo is not None else ""

            data_parsed = self._parse_data(data_raw)
            if data_parsed:
                ultima_data = data_parsed

            # --- voci scartate per tipo ---
            if not descr and not tipo:
                continue
            if tipo in TIPO_ESCLUSI:
                skip_buckets["p (personali)"] = skip_buckets.get("p (personali)", 0) + 1
                continue
            if tipo in TIPO_FASE2:
                skip_buckets[f"{tipo} (fase 2)"] = skip_buckets.get(f"{tipo} (fase 2)", 0) + 1
                continue

            is_imu = tipo == "g" and "imu" in descr.lower()
            if tipo == "g" and not is_imu:
                skip_buckets["g (giroconto, escluso)"] = skip_buckets.get("g (giroconto, escluso)", 0) + 1
                continue
            if tipo not in ("a",) and not is_imu:
                skip_buckets[f"{tipo or '(vuoto)'} (non gestito)"] = (
                    skip_buckets.get(f"{tipo or '(vuoto)'} (non gestito)", 0) + 1
                )
                continue

            # --- individua importo + proprietario ---
            uscite = [(idx, v) for idx in COL_USCITA_OWNER
                      if (v := self._num(row[idx]))]
            entrate = [v for idx in COL_ENTRATA if (v := self._num(row[idx]))]
            if not uscite:
                if entrate:
                    warnings.append(
                        f"riga {n}: «{descr}» è un'ENTRATA ({sum(entrate):g}€), "
                        f"non una spesa — saltata (verificare a mano)"
                    )
                else:
                    warnings.append(f"riga {n}: «{descr}» tipo {tipo} senza importo — saltata")
                continue
            if len(uscite) > 1:
                warnings.append(
                    f"riga {n}: «{descr}» ha importi in più colonne proprietario — saltata"
                )
                continue
            col_idx, importo_raw = uscite[0]
            importo = Decimal(str(importo_raw)).quantize(Decimal("0.01"))
            owner = owners[COL_USCITA_OWNER[col_idx]]

            # --- data ---
            data_eff = data_parsed or ultima_data
            data_stimata = data_parsed is None
            if data_eff is None:
                data_eff = dt.date(2022, 10, 1)
            note_extra = " · DATA STIMATA" if data_stimata else ""
            if data_stimata:
                warnings.append(f"riga {n}: «{descr}» senza data nel foglio → stimata {data_eff}")

            # --- IMU: anticipata da Bruna per conto di Fabio ---
            riferimento_quota = None
            if is_imu:
                categoria = cat_imu
                m = re.search(r"\b(fabio|sandro|bruna)\b", descr.lower())
                if m:
                    kw = {"fabio": "Fabio", "sandro": "Alessandro", "bruna": "Bruna"}[m.group(1)]
                    rif = owners[kw]
                    if rif != owner:
                        riferimento_quota = rif
            else:
                categoria = cat_arredo

            marker = f"[conti2023#{n}]"
            note = f"Importato da Contabilità.xlsx «conti 2023» riga {n}.{note_extra} {marker}"
            descrizione = descr[:300] or f"(riga {n})"

            if dry_run:
                self.stdout.write(
                    f"  riga {n:>3} | {data_eff} | {owner_kw(owner):<10} -{importo:>8}€ | "
                    f"{categoria.codice:<18} | {descrizione[:40]}"
                )
                creati += 1
                tot_importato += importo
                continue

            obj = Expense.objects.filter(note__contains=marker).first()
            defaults = dict(
                data=data_eff,
                category=categoria,
                importo=importo,
                descrizione=descrizione,
                anticipata_da_owner=owner,
                ripartibile_su_inquilini=False,
                riferimento_quota_owner=riferimento_quota,
                note=note,
            )
            if obj:
                for k, v in defaults.items():
                    setattr(obj, k, v)
                obj.save()
                aggiornati += 1
            else:
                Expense.objects.create(**defaults)
                creati += 1
            tot_importato += importo

        return {
            "creati": creati,
            "aggiornati": aggiornati,
            "tot_importato": tot_importato,
            "skip_buckets": skip_buckets,
            "warnings": warnings,
        }

    # ------------------------------------------------------------------
    @staticmethod
    def _categoria(codice, nome, dry_run):
        from billing.models import ExpenseCategory

        cat = ExpenseCategory.objects.filter(codice=codice).first()
        if cat:
            return cat
        if dry_run:
            return ExpenseCategory(codice=codice, nome=nome, ripartibile_inquilini=False)
        return ExpenseCategory.objects.create(
            codice=codice, nome=nome, ripartibile_inquilini=False
        )

    @staticmethod
    def _owner(keyword):
        from properties.models import OwnerProfile

        for o in OwnerProfile.objects.all():
            if keyword.lower() in str(o).lower():
                return o
        raise CommandError(f"OwnerProfile per «{keyword}» non trovato.")

    # ------------------------------------------------------------------
    # Fase 2 — affitti mancanti + correzione cauzioni
    # ------------------------------------------------------------------
    def _fase2(self, *, dry_run: bool) -> dict:
        import calendar
        from datetime import date

        from django.db.models import Q

        from billing.models import Receivable
        from properties.models import RoomAssignment, TenantProfile

        aff_rows: list[dict] = []
        corr_rows: list[dict] = []
        cau_rows: list[dict] = []
        warnings: list[str] = []

        # --- affitti ---
        for tenant_key, anno, mese, dovuto, fonte, nota in AFFITTO_RICOSTRUITO:
            tenant = TenantProfile.objects.filter(nominativo__icontains=tenant_key).first()
            if tenant is None:
                warnings.append(f"affitto {tenant_key} {anno}-{mese:02}: inquilino non trovato")
                continue
            comp_da = date(anno, mese, 1)
            comp_a = date(anno, mese, calendar.monthrange(anno, mese)[1])
            scadenza = date(anno, mese, 6)
            assignment = (
                RoomAssignment.objects.filter(tenant=tenant, valid_from__lte=comp_a)
                .filter(Q(valid_to__isnull=True) | Q(valid_to__gte=comp_da))
                .order_by("-valid_from").first()
            )
            if assignment is None:
                warnings.append(
                    f"affitto {tenant_key} {anno}-{mese:02}: RoomAssignment mancante"
                )
                continue
            importo = Decimal(dovuto).quantize(Decimal("0.01"))
            existing = Receivable.objects.filter(
                causale="affitto", assignment=assignment,
                competenza_da=comp_da, competenza_a=comp_a,
            ).first()

            if existing and MARKER_AFFITTO not in (existing.note or ""):
                stato = "GIÀ PRESENTE (non importato da me) — non tocco"
            else:
                stato = "aggiorno" if existing else "creo"
                if not dry_run:
                    Receivable.objects.update_or_create(
                        causale="affitto", assignment=assignment,
                        competenza_da=comp_da, competenza_a=comp_a,
                        defaults=dict(
                            scadenza=scadenza,
                            importo_dovuto=importo,
                            note=f"Ricostruito da Contabilità.xlsx «conti 2023» "
                                 f"({fonte}). {nota} {MARKER_AFFITTO}",
                        ),
                    )
            aff_rows.append(dict(
                tenant=tenant_key, anno=anno, mese=mese, importo=importo,
                fonte=fonte, nota=nota, stato=stato,
            ))

        # --- correzione affitti già a sistema ---
        for tenant_key, anno, mese, importo_str, nota in AFFITTO_CORREZIONI:
            comp_da = date(anno, mese, 1)
            rec = Receivable.objects.filter(
                causale="affitto",
                assignment__tenant__nominativo__icontains=tenant_key,
                competenza_da=comp_da,
            ).first()
            if rec is None:
                warnings.append(
                    f"correzione affitto {tenant_key} {anno}-{mese:02}: Receivable non trovato"
                )
                continue
            nuovo = Decimal(importo_str).quantize(Decimal("0.01"))
            vecchio = rec.importo_dovuto
            corr_rows.append(dict(
                tenant=tenant_key, anno=anno, mese=mese,
                vecchio=vecchio, nuovo=nuovo, nota=nota, cambia=(vecchio != nuovo),
            ))
            if not dry_run and vecchio != nuovo:
                rec.importo_dovuto = nuovo
                if MARKER_AFFITTO not in (rec.note or ""):
                    rec.note = (rec.note + " " if rec.note else "") + (
                        f"Importo corretto da Contabilità.xlsx «conti 2023» "
                        f"({nota}); era {vecchio}. {MARKER_AFFITTO}"
                    )
                rec.save(update_fields=["importo_dovuto", "note"])

        # --- cauzioni ---
        for tenant_key, importo_str, nota in CAUZIONI_CORREZIONI:
            dep = (
                Receivable.objects.filter(
                    causale="deposito",
                    assignment__tenant__nominativo__icontains=tenant_key,
                    competenza_da__year__lte=2023,
                    importo_dovuto__gt=0,
                )
                .order_by("competenza_da").first()
            )
            if dep is None:
                warnings.append(f"cauzione {tenant_key}: Receivable deposito 2023 non trovato")
                continue
            nuovo = Decimal(importo_str).quantize(Decimal("0.01"))
            vecchio = dep.importo_dovuto
            cau_rows.append(dict(
                tenant=tenant_key, vecchio=vecchio, nuovo=nuovo, nota=nota,
                cambia=(vecchio != nuovo),
            ))
            if not dry_run and vecchio != nuovo:
                dep.importo_dovuto = nuovo
                if MARKER_CAUZIONE not in (dep.note or ""):
                    dep.note = (dep.note + " " if dep.note else "") + (
                        f"Importo corretto da Contabilità.xlsx «conti 2023» "
                        f"({nota}); era {vecchio}. {MARKER_CAUZIONE}"
                    )
                dep.save(update_fields=["importo_dovuto", "note"])

        return {"aff_rows": aff_rows, "corr_rows": corr_rows,
                "cau_rows": cau_rows, "warnings": warnings}

    # ------------------------------------------------------------------
    # Fase 3 — riconciliazione storica: BT sintetiche dal libro mano
    # ------------------------------------------------------------------
    def _fase3(self, *, dry_run: bool) -> dict:
        from datetime import date

        from billing.models import Receivable

        aff_rows: list[dict] = []
        dep_rows: list[dict] = []
        ute_rows: list[dict] = []
        corr_rows: list[dict] = []
        warnings: list[str] = []

        # --- affitti ---
        for tenant_key, anno, mese, tranches, fonte, nota in AFFITTO_FASE3:
            rec = Receivable.objects.filter(
                causale="affitto",
                assignment__tenant__nominativo__icontains=tenant_key,
                competenza_da=date(anno, mese, 1),
            ).first()
            if rec is None:
                warnings.append(
                    f"affitto {tenant_key} {anno}-{mese:02}: Receivable non trovato"
                )
                continue
            key = f"affitto:{tenant_key}:{anno}-{mese:02}"
            row = self._riconcilia_da_foglio(
                rec, tranches, fonte, nota, "Affitto", key, dry_run
            )
            row.update(tenant=tenant_key, anno=anno, mese=mese)
            aff_rows.append(row)

        # --- depositi versati ---
        for tenant_key, tranches, fonte, nota in DEPOSITO_FASE3:
            rec = (
                Receivable.objects.filter(
                    causale="deposito",
                    assignment__tenant__nominativo__icontains=tenant_key,
                    importo_dovuto__gt=0,
                    competenza_da__year__in=(2022, 2023),
                )
                .order_by("competenza_da")
                .first()
            )
            if rec is None:
                warnings.append(f"deposito {tenant_key}: Receivable non trovato")
                continue
            key = f"deposito:{tenant_key}"
            row = self._riconcilia_da_foglio(
                rec, tranches, fonte, nota, "Cauzione", key, dry_run
            )
            row.update(tenant=tenant_key)
            dep_rows.append(row)

        # --- utenze 2023 pagate cash (libro mano) ---
        for tenant_key, anno, mese, tranches, fonte, nota in UTENZE_FASE3:
            rec = Receivable.objects.filter(
                causale="utenze",
                assignment__tenant__nominativo__icontains=tenant_key,
                competenza_da=date(anno, mese, 1),
            ).first()
            if rec is None:
                warnings.append(
                    f"utenze {tenant_key} {anno}-{mese:02}: Receivable non trovato"
                )
                continue
            key = f"utenze:{tenant_key}:{anno}-{mese:02}"
            row = self._riconcilia_da_foglio(
                rec, tranches, fonte, nota, "Utenze", key, dry_run
            )
            row.update(tenant=tenant_key, anno=anno, mese=mese)
            ute_rows.append(row)

        # --- correzioni rate condominio (+1€, vince il foglio) ---
        from billing.models import Expense

        for descr_esatta, vecchio_str, nuovo_str, fonte in CORREZIONI_CONDOMINIO:
            match = Expense.objects.filter(
                category__codice="condominio", descrizione=descr_esatta
            )
            if match.count() != 1:
                warnings.append(
                    f"correzione condominio «{descr_esatta}»: "
                    f"{match.count()} Expense trovate (attesa 1) — salto"
                )
                continue
            exp = match.first()
            vecchio = exp.importo
            nuovo = Decimal(nuovo_str)
            atteso = Decimal(vecchio_str)
            row = dict(descr=exp.descrizione, vecchio=vecchio, nuovo=nuovo, fonte=fonte)
            if vecchio == nuovo:
                row["stato"] = "già corretto"
            elif vecchio != atteso:
                row["stato"] = f"⚠ importo inatteso ({vecchio}) — non modifico"
            else:
                row["stato"] = "corretto"
                if not dry_run:
                    exp.importo = nuovo
                    if MARKER_CORR_COND not in (exp.note or ""):
                        exp.note = ((exp.note + " ") if exp.note else "") + (
                            f"Importo corretto da Contabilità.xlsx «conti 2023» "
                            f"({fonte}); era {vecchio}. {MARKER_CORR_COND}"
                        )
                    exp.save(update_fields=["importo", "note", "updated_at"])
            corr_rows.append(row)

        return {
            "aff_rows": aff_rows,
            "dep_rows": dep_rows,
            "ute_rows": ute_rows,
            "corr_rows": corr_rows,
            "warnings": warnings,
        }

    def _riconcilia_da_foglio(
        self, rec, tranches, fonte, nota, label, key, dry_run
    ) -> dict:
        """Sintetizza le BankTransaction del libro mano per un Receivable e le
        alloca. Ritorna un dict descrittivo per il report."""
        from datetime import date

        from billing.models import BankTransaction, BankTransactionAllocation

        dovuto = rec.importo_dovuto
        info = dict(dovuto=dovuto, fonte=fonte, nota=nota, importo_bt=None)

        # Già riconciliato da un movimento bancario reale (non sintetico): non
        # tocco — l'estratto conto vince sul libro mano.
        if rec.allocations.exclude(
            bank_transaction__note__contains=MARKER_BT
        ).exists():
            info["stato"] = "già riconciliato da movimento reale — salto"
            return info

        if not tranches:
            info["stato"] = "nessun movimento nel foglio — resta atteso"
            info["importo_bt"] = Decimal("0")
            return info

        tenant_nom = rec.assignment.tenant.nominativo
        importo_tot = sum((Decimal(t[2]) for t in tranches), Decimal("0"))
        info["importo_bt"] = importo_tot

        if dry_run:
            if importo_tot + Decimal("1") >= dovuto:
                info["stato"] = "creerò BT + allocazione → pagato"
            else:
                info["stato"] = (
                    f"creerò BT parziale → scostamento (manca "
                    f"{dovuto - importo_tot}€)"
                )
            return info

        residuo = dovuto
        for i, (owner_key, data_iso, importo_str) in enumerate(tranches):
            account = self._owner_account(owner_key)
            importo = Decimal(importo_str)
            marker = f"{MARKER_BT} {key}#{i}]"
            descrizione = (
                f"{label} {tenant_nom} — da libro mano «conti 2023» ({fonte})"
            )
            note = (
                f"Movimento ricostruito dal libro mano Contabilità.xlsx "
                f"«conti 2023» ({fonte}). {nota} {marker}"
            ).strip()
            bt = BankTransaction.objects.filter(note__contains=marker).first()
            if bt:
                bt.data = date.fromisoformat(data_iso)
                bt.importo = importo
                bt.descrizione = descrizione
                bt.owner_account = account
                bt.note = note
                bt.save()
            else:
                bt = BankTransaction.objects.create(
                    data=date.fromisoformat(data_iso),
                    importo=importo,
                    descrizione=descrizione,
                    owner_account=account,
                    note=note,
                )
            quota = min(importo, residuo)
            if quota > 0:
                BankTransactionAllocation.objects.update_or_create(
                    bank_transaction=bt,
                    receivable=rec,
                    defaults={"importo": quota},
                )
                residuo -= quota

        rec.refresh_from_db()
        if rec.stato == "pagato":
            extra = importo_tot - dovuto
            info["stato"] = "✓ pagato" + (
                f" (residuo bonifico {extra}€)" if extra > 0 else ""
            )
        else:
            info["stato"] = (
                f"⚠ scostamento: foglio {importo_tot}€ < dovuto {dovuto}€ "
                f"(manca {dovuto - importo_tot}€) — resta atteso"
            )
        return info

    def _owner_account(self, keyword):
        """OwnerBankAccount del proprietario individuato da ``keyword``."""
        from properties.models import OwnerBankAccount

        owner = self._owner(keyword)
        acc = OwnerBankAccount.objects.filter(owner=owner).first()
        if acc is None:
            raise CommandError(f"OwnerBankAccount per «{keyword}» non trovato.")
        return acc

    # ------------------------------------------------------------------
    # Fase 4 — bollette utenze (PDF Edison → UtilityBill, TARI → Expense)
    # ------------------------------------------------------------------
    def _fase4(self, *, bollette_dir: str, dry_run: bool) -> dict:
        from django.core.files import File
        from django.db.models import Q

        from billing.models import Supplier, UtilityBill

        pdf_rows: list[dict] = []
        warnings: list[str] = []
        bdir = Path(bollette_dir).expanduser().resolve()
        sandro = self._owner("Alessandro")

        supplier = Supplier.objects.filter(nome__iexact="Edison Energia").first()
        if supplier is None and not dry_run:
            supplier = Supplier.objects.create(
                nome="Edison Energia", tipo=Supplier.TipoFornitore.ENERGIA
            )

        pdfs = sorted(bdir.glob("utenze-*edison.pdf")) if bdir.is_dir() else []
        if not pdfs:
            warnings.append(f"nessun PDF «utenze-*edison.pdf» in {bdir}")

        for path in pdfs:
            meta = self._parse_bolletta_edison(path)
            if meta is None:
                warnings.append(f"{path.name}: nome file non interpretabile — salto")
                continue
            stem = path.stem

            # Idempotenza: numero_fattura = nome file. Se non c'è, cerca una
            # UtilityBill già a sistema per la stessa bolletta (stesso prodotto,
            # anno e importo ~uguale) — sono le #144/#145/#147 importate prima
            # con metadati grezzi: vanno aggiornate sul posto, non duplicate.
            bill = UtilityBill.objects.filter(numero_fattura=stem).first()
            if bill is None:
                simili = [
                    b
                    for b in UtilityBill.objects.filter(
                        prodotto=meta["prodotto"]
                    ).filter(
                        Q(periodo_da__year=meta["anno"])
                        | Q(periodo_a__year=meta["anno"])
                    )
                    if abs(b.importo_totale - meta["importo_file"]) < Decimal("1")
                ]
                if len(simili) == 1:
                    bill = simili[0]
                elif len(simili) > 1:
                    warnings.append(
                        f"{stem}: {len(simili)} UtilityBill simili a sistema "
                        f"(ID {[b.id for b in simili]}) — ne creo una nuova, verificare"
                    )

            # Importo: se il PDF non ha dato i centesimi ma la UB esistente sì
            # (es. #144 = 51,61), conserva il valore più preciso già a sistema.
            # Il nome file porta la parte intera dell'importo (51 ← 51,61).
            importo = meta["importo"]
            if (
                bill is not None
                and importo == meta["importo_file"]
                and bill.importo_totale
                and int(bill.importo_totale) == int(meta["importo_file"])
            ):
                importo = bill.importo_totale

            azione = "aggiornata" if bill is not None else "creata"
            row = dict(
                nome=stem, prodotto=meta["prodotto"], periodo_da=meta["periodo_da"],
                periodo_a=meta["periodo_a"], importo=importo,
                importo_file=meta["importo_file"], fonte=meta["fonte"],
                azione=azione, pre_id=(bill.id if bill is not None else None),
            )
            if dry_run:
                pdf_rows.append(row)
                continue

            if bill is None:
                bill = UtilityBill(numero_fattura=stem)
            bill.numero_fattura = stem
            bill.supplier = supplier
            bill.prodotto = meta["prodotto"]
            bill.data_emissione = meta["data_emissione"]
            bill.periodo_da = meta["periodo_da"]
            bill.periodo_a = meta["periodo_a"]
            bill.importo_totale = importo
            bill.pagata_da_owner = sandro
            with path.open("rb") as fh:
                bill.file_pdf.save(path.name, File(fh), save=False)
            # Un solo save() → un solo post_save → una sola Expense generata.
            bill.save()
            row["pre_id"] = bill.id
            pdf_rows.append(row)

        foglio_rows = self._importa_bollette_foglio(dry_run)
        tari = self._importa_tari(dry_run)
        return {"pdf_rows": pdf_rows, "foglio_rows": foglio_rows, "tari": tari,
                "warnings": warnings, "bdir": str(bdir)}

    def _importa_bollette_foglio(self, dry_run: bool) -> list[dict]:
        """Bollette pre-voltura registrate solo nel libro mano (nessun PDF):
        UtilityBill senza `pagata_da_owner`, quindi senza Expense generata."""
        from billing.models import Supplier, UtilityBill

        rows: list[dict] = []
        supplier = None
        if not dry_run:
            supplier, _ = Supplier.objects.get_or_create(
                nome="Sconosciuto",
                defaults={"tipo": Supplier.TipoFornitore.ALTRO},
            )
        for riga, pda, pa, dpag, imp, descr in BOLLETTE_FOGLIO:
            nf = f"conti2023-bolletta-riga{riga}"
            importo = Decimal(imp)
            periodo_da = dt.date.fromisoformat(pda)
            periodo_a = dt.date.fromisoformat(pa)
            bill = UtilityBill.objects.filter(numero_fattura=nf).first()
            row = dict(
                riga=riga, periodo_da=periodo_da, periodo_a=periodo_a,
                importo=importo, descr=descr,
                stato=("aggiornata" if bill else "creata"),
            )
            if dry_run:
                rows.append(row)
                continue
            note = (
                f"Bolletta aggregata pagata da Bruna prima della voltura del "
                f"contratto utenze a nome di Sandro: fattura non intestata a "
                f"lui, nessun PDF. Importo dal libro mano «conti 2023» riga "
                f"{riga} «{descr}», pagata il {dpag}. Prodotto (luce/gas) non "
                f"determinabile dal libro mano."
            )
            if bill is None:
                bill = UtilityBill(numero_fattura=nf)
            bill.numero_fattura = nf
            bill.supplier = supplier
            bill.prodotto = UtilityBill.Prodotto.LUCE  # segnaposto: vedi note
            bill.data_emissione = periodo_a
            bill.periodo_da = periodo_da
            bill.periodo_a = periodo_a
            bill.importo_totale = importo
            bill.pagata_da_owner = None  # nessuna Expense generata dal signal
            bill.note = note
            bill.save()
            rows.append(row)
        return rows

    def _parse_bolletta_edison(self, path: Path) -> dict | None:
        """Metadati di una bolletta Edison: dal nome file (sempre) arricchiti
        dal testo del PDF quando leggibile (i PDF 2022 sono scansioni)."""
        import subprocess

        name = path.name
        m_prod = re.search(r"(luce|gas)", name, re.IGNORECASE)
        m_date = re.search(r"(\d{4})-(\d{2})-+(\d{1,4})", name)
        if not (m_prod and m_date):
            return None
        prodotto = "luce" if m_prod.group(1).lower() == "luce" else "gas"
        anno, mese = int(m_date.group(1)), int(m_date.group(2))
        importo = Decimal(m_date.group(3))
        importo_file = importo
        periodo_da = dt.date(anno, mese, 1)
        periodo_a = dt.date(anno, mese, calendar.monthrange(anno, mese)[1])
        data_emissione = periodo_a
        fonte = "nome file"

        try:
            txt = subprocess.run(
                ["pdftotext", "-layout", str(path), "-"],
                capture_output=True, text=True, timeout=30,
            ).stdout
        except (subprocess.SubprocessError, FileNotFoundError, OSError):
            txt = ""

        m = re.search(
            r"MERCATO LIBERO\s+(\w+)\s+(\d{4})\s*-\s*(\w+)\s+(\d{4})", txt
        )
        if m:
            m1 = MESI_IT.get(m.group(1).lower())
            m2 = MESI_IT.get(m.group(3).lower())
            if m1 and m2:
                periodo_da = dt.date(int(m.group(2)), m1, 1)
                periodo_a = dt.date(
                    int(m.group(4)), m2, calendar.monthrange(int(m.group(4)), m2)[1]
                )
                fonte = "PDF"
        m = re.search(
            r"Bolletta N°\s*\d+\s*del\s+(\d{1,2})\s+(\w+)\s+(\d{4})", txt
        )
        if m:
            mm = MESI_IT.get(m.group(2).lower())
            if mm:
                data_emissione = dt.date(int(m.group(3)), mm, int(m.group(1)))
        m = re.search(r"([\d.]+,\d{2})\s+Totale bolletta", txt) or re.search(
            r"Totale bolletta\s+([\d.]+,\d{2})", txt
        )
        if m:
            val = Decimal(m.group(1).replace(".", "").replace(",", "."))
            if abs(val - importo_file) <= 2:  # sanity-check contro il nome file
                importo = val

        return dict(
            prodotto=prodotto, anno=anno, periodo_da=periodo_da, periodo_a=periodo_a,
            data_emissione=data_emissione, importo=importo,
            importo_file=importo_file, fonte=fonte,
        )

    def _importa_tari(self, dry_run: bool) -> dict:
        """La TARI 2023 (riga 121 del foglio) come Expense categoria «tari»."""
        from billing.models import Expense, ExpenseCategory

        importo = Decimal("405")
        bruna = self._owner("Bruna")
        exp = Expense.objects.filter(note__contains=MARKER_TARI).first()
        row = dict(importo=importo, stato="aggiornerò" if exp else "creerò")
        if dry_run:
            return row

        cat = ExpenseCategory.objects.filter(codice="tari").first()
        if cat is None:
            cat = ExpenseCategory.objects.create(
                codice="tari", nome="TARI", ripartibile_inquilini=True
            )
        defaults = dict(
            data=dt.date(2023, 7, 1),
            category=cat,
            importo=importo,
            descrizione="TARI 2023 — da libro mano «conti 2023» riga 121",
            anticipata_da_owner=bruna,
            ripartibile_su_inquilini=False,
            note=f"Importato da Contabilità.xlsx «conti 2023» riga 121. {MARKER_TARI}",
        )
        if exp:
            for k, v in defaults.items():
                setattr(exp, k, v)
            exp.save()
            row["stato"] = "aggiornata"
        else:
            Expense.objects.create(**defaults)
            row["stato"] = "creata"
        return row

    def _report_fase4(self, s: dict):
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(
            "=== Fase 4 — bollette utenze (Edison) ==="))
        self.stdout.write(f"  cartella PDF: {s['bdir']}")
        tot = Decimal("0")
        for r in s["pdf_rows"]:
            tot += r["importo"]
            avviso = ""
            if r["importo"] != r["importo_file"]:
                avviso = f" (nome file: {r['importo_file']})"
            etichetta = r["azione"]
            if r["azione"] == "aggiornata" and r["pre_id"] is not None:
                etichetta = f"aggiornata #{r['pre_id']}"
            self.stdout.write(
                f"  {etichetta:<16} {r['prodotto']:<5} "
                f"{r['periodo_da']} → {r['periodo_a']}  {r['importo']:>8}€"
                f"{avviso}  [periodo da {r['fonte']}]  {r['nome']}"
            )
        self.stdout.write(f"  → {len(s['pdf_rows'])} bollette da PDF, totale {tot}€")

        if s.get("foglio_rows"):
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(
                "=== Fase 4 — bollette dal libro mano (pre-voltura, senza PDF) ==="))
            tot_f = Decimal("0")
            for r in s["foglio_rows"]:
                tot_f += r["importo"]
                self.stdout.write(
                    f"  {r['stato']:<11} riga {r['riga']:<4} "
                    f"{r['periodo_da']} → {r['periodo_a']}  {r['importo']:>8}€  "
                    f"«{r['descr']}»"
                )
            self.stdout.write(
                f"  → {len(s['foglio_rows'])} UtilityBill senza pagata_da_owner "
                f"(niente Expense), totale {tot_f}€"
            )

        t = s["tari"]
        self.stdout.write("")
        self.stdout.write(
            f"  TARI 2023 (libro mano riga 121): {t['importo']}€  [{t['stato']}] "
            f"— affianca la «TARI 2023 (F24)» già a sistema"
        )
        for w in s["warnings"]:
            self.stdout.write(self.style.WARNING(f"  ⚠ {w}"))

    # ------------------------------------------------------------------
    @staticmethod
    def _num(v):
        """Importo della cella: numero o stringa numerica (anche con virgola)."""
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            return float(v) or None
        if isinstance(v, str):
            s = v.strip().replace(".", "").replace(",", ".") if "," in v else v.strip()
            try:
                return float(s) or None
            except ValueError:
                return None
        return None

    @staticmethod
    def _parse_data(d):
        """Parsing best-effort delle date del foglio (formati eterogenei)."""
        if isinstance(d, dt.datetime):
            return d.date()
        if isinstance(d, dt.date):
            return d
        if d is None:
            return None
        s = str(d).strip()
        # gg/mm/aaaa (anche con doppi separatori e refusi: 25/1/0123, 4/7//23)
        m = re.match(r"(\d{1,2})/+(\d{1,2})/+(\d{2,4})", s)
        if m:
            g, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3)[-2:])
            try:
                return dt.date(2000 + y, mo, g)
            except ValueError:
                return None
        # gg:mm:aa — data digitata come orario (01:11:22, 05:01:23)
        m = re.match(r"(\d{2}):(\d{2}):(\d{2})$", s)
        if m:
            g, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            try:
                return dt.date(2000 + y, mo, g)
            except ValueError:
                return None
        return None

    # ------------------------------------------------------------------
    def _report(self, s: dict):
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("=== Import spese proprietari «conti 2023» ==="))
        self.stdout.write(
            f"  Expense create: {s['creati']}, aggiornate: {s['aggiornati']}, "
            f"totale importato: {s['tot_importato']}€"
        )
        if s["skip_buckets"]:
            self.stdout.write("  Righe saltate (per fase/tipo):")
            for k, v in sorted(s["skip_buckets"].items()):
                self.stdout.write(f"      {v:>3} × {k}")
        for w in s["warnings"]:
            self.stdout.write(self.style.WARNING(f"  ⚠ {w}"))

    _MESI = ("", "gen", "feb", "mar", "apr", "mag", "giu",
             "lug", "ago", "set", "ott", "nov", "dic")

    def _report_fase2(self, s: dict):
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(
            "=== Fase 2 — affitti ricostruiti «conti 2023» ==="))
        tenant_corrente = None
        tot = Decimal("0")
        for r in s["aff_rows"]:
            if r["tenant"] != tenant_corrente:
                tenant_corrente = r["tenant"]
                self.stdout.write(f"  {tenant_corrente}:")
            mese = f"{self._MESI[r['mese']]} {r['anno']}"
            self.stdout.write(
                f"      {mese:<9} {r['importo']:>8}€  [{r['stato']:<38}] "
                f"{r['fonte']:<16} — {r['nota']}"
            )
            if "creo" in r["stato"] or "aggiorno" in r["stato"]:
                tot += r["importo"]
        self.stdout.write(f"  Totale affitti creati/aggiornati: {tot}€")

        if s["corr_rows"]:
            self.stdout.write("")
            self.stdout.write(self.style.MIGRATE_HEADING(
                "=== Fase 2 — correzione affitti già a sistema ==="))
            for r in s["corr_rows"]:
                mese = f"{self._MESI[r['mese']]} {r['anno']}"
                freccia = (f"{r['vecchio']} → {r['nuovo']}" if r["cambia"]
                           else f"{r['nuovo']} (invariato)")
                self.stdout.write(f"  {r['tenant']:<14} {mese:<9} {freccia:<18} — {r['nota']}")

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("=== Fase 2 — cauzioni ==="))
        for r in s["cau_rows"]:
            freccia = f"{r['vecchio']} → {r['nuovo']}" if r["cambia"] else f"{r['nuovo']} (invariata)"
            self.stdout.write(f"  {r['tenant']:<10} {freccia:<20} — {r['nota']}")

        for w in s["warnings"]:
            self.stdout.write(self.style.WARNING(f"  ⚠ {w}"))

    def _report_fase3(self, s: dict):
        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(
            "=== Fase 3 — riconciliazione affitti dal libro mano «conti 2023» ==="))
        tenant_corrente = None
        n_pag = n_scost = n_skip = 0
        for r in s["aff_rows"]:
            if r["tenant"] != tenant_corrente:
                tenant_corrente = r["tenant"]
                self.stdout.write(f"  {tenant_corrente}:")
            mese = f"{self._MESI[r['mese']]} {r['anno']}"
            imp = f"{r['importo_bt']}€" if r["importo_bt"] is not None else "—"
            self.stdout.write(
                f"      {mese:<9} dovuto {r['dovuto']:>8}€  foglio {imp:<9} "
                f"[{r['stato']}]"
            )
            if "pagato" in r["stato"]:
                n_pag += 1
            elif "salto" in r["stato"]:
                n_skip += 1
            else:
                n_scost += 1
        self.stdout.write(
            f"  → affitti: {n_pag} riconciliati, {n_scost} da verificare, "
            f"{n_skip} saltati (già a posto)"
        )

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING("=== Fase 3 — cauzioni ==="))
        for r in s["dep_rows"]:
            imp = f"{r['importo_bt']}€" if r["importo_bt"] is not None else "—"
            self.stdout.write(
                f"  {r['tenant']:<10} dovuto {r['dovuto']:>8}€  foglio {imp:<9} "
                f"[{r['stato']}]"
            )

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(
            "=== Fase 3 — utenze cash dal libro mano «conti 2023» ==="))
        tenant_corrente = None
        n_pag = n_scost = n_skip = 0
        for r in s.get("ute_rows", []):
            if r["tenant"] != tenant_corrente:
                tenant_corrente = r["tenant"]
                self.stdout.write(f"  {tenant_corrente}:")
            mese = f"{self._MESI[r['mese']]} {r['anno']}"
            imp = f"{r['importo_bt']}€" if r["importo_bt"] is not None else "—"
            self.stdout.write(
                f"      {mese:<9} dovuto {r['dovuto']:>8}€  foglio {imp:<9} "
                f"[{r['stato']}]"
            )
            if "pagato" in r["stato"]:
                n_pag += 1
            elif "salto" in r["stato"]:
                n_skip += 1
            else:
                n_scost += 1
        if s.get("ute_rows"):
            self.stdout.write(
                f"  → utenze: {n_pag} riconciliate, {n_scost} da verificare, "
                f"{n_skip} saltate (già a posto)"
            )

        self.stdout.write("")
        self.stdout.write(self.style.MIGRATE_HEADING(
            "=== Fase 3 — correzioni rate condominio (vince il foglio) ==="))
        for r in s["corr_rows"]:
            freccia = f"{r['vecchio']} → {r['nuovo']}"
            self.stdout.write(
                f"  {r['descr'][:44]:<44} {freccia:<16} [{r['stato']}]"
            )

        for w in s["warnings"]:
            self.stdout.write(self.style.WARNING(f"  ⚠ {w}"))


def owner_kw(owner):
    s = str(owner)
    return s.split()[0]
